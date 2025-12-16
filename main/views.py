from django.shortcuts import render, redirect, HttpResponse, get_object_or_404
from django.http import JsonResponse
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required
from main.models import *
from django.contrib import messages
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
import json
import re
import os
from uuid import uuid4
import aiohttp
from datetime import datetime, timedelta
from django.db.models import Sum, Q, Max, F, Count
from django.utils import timezone
from django.utils.text import slugify
from io import BytesIO
from xhtml2pdf import pisa
from weasyprint import HTML
from weasyprint.text.fonts import FontConfiguration
from django.template.loader import render_to_string
from functools import wraps
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

import json
from django.db.models import Sum
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.db import IntegrityError

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from django.urls import reverse
from urllib.parse import urlencode, unquote
from decimal import Decimal

import logging
logger = logging.getLogger(__name__)

def ceoadmin_required(view_func):
    @wraps(view_func)
    def _wrapped_view(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect("login")
        if request.user.username != "ceoadmin":
            messages.warning(request, "Sizda bu bo'limga kirish huquqi yo'q.")
            return redirect("home")
        return view_func(request, *args, **kwargs)
    return _wrapped_view


# Custom decorator for user authentication and username check

def build_contract_filter_params(request):
    """Build filter parameters from request for contracts"""
    filters = {}
    params_dict = {}
    
    city_id = request.GET.get("city")
    building_id = request.GET.get("building")
    debt_status = request.GET.get("debt")
    status = request.GET.get("status")
    q = request.GET.get("q")
    
    if city_id and city_id.isdigit():
        filters["home__building__city__id"] = city_id
        params_dict["city"] = city_id
        
    if building_id and building_id.isdigit():
        filters["home__building__id"] = building_id
        params_dict["building"] = building_id
        
    if debt_status:
        filters['debt'] = debt_status == "true"
        params_dict["debt"] = debt_status
        
    if status and status in ["0", "1", "2", "3"]:
        status_map = {
            "0": "Bekor qilingan",
            "1": "Rasmiylashtirilmoqda", 
            "2": "Rasmiylashtirilgan",
            "3": "Tugallangan",
        }
        filters["status"] = status_map[status]
        params_dict["status"] = status
        
    if q:
        params_dict["q"] = q
    
    return filters, params_dict

def build_contract_filter_url(params_dict):
    """Build filter URL from parameters dictionary for contracts"""
    if not params_dict:
        return ""
    
    url_parts = []
    for key, value in params_dict.items():
        url_parts.append(f"{key}={value}")
    
    return "?" + "&".join(url_parts)

def build_contract_next_url(request):
    """Build next URL with current filters for contracts"""
    filters, params_dict = build_contract_filter_params(request)
    base_url = reverse('contract')
    if params_dict:
        return f"{base_url}?{urlencode(params_dict)}"
    return base_url

def get_contract_next_url(request):
    """Get the next URL from request parameters for contracts"""
    next_url = request.GET.get('next', '')
    if next_url:
        return unquote(next_url)
    return reverse('contract')


def tushum_view(request):
    today = timezone.now().date()
    one_week_ago = today - timedelta(days=7)
    one_month_ago = today - timedelta(days=30)

    # Konsalting shartnomalari bo'yicha tushumlarni hisoblash
    payments = ConsultingContract.objects.all()

    def _sum(qs):
        return qs.aggregate(total=Sum('amount_paid'))['total'] or 0

    kunlik_tushum = _sum(payments.filter(updated_at__date=today))
    haftalik_tushum = _sum(payments.filter(updated_at__date__gte=one_week_ago))
    oylik_tushum = _sum(payments.filter(updated_at__date__gte=one_month_ago))
    umumiy = _sum(payments)

    return {
        "status": payments.exists(),
        "kunlik_tushum": "{:,}".format(int(kunlik_tushum)).replace(",", " "),
        "haftalik_tushum": "{:,}".format(int(haftalik_tushum)).replace(",", " "),
        "oylik_tushum": "{:,}".format(int(oylik_tushum)).replace(",", " "),
        "umumiy": "{:,}".format(int(umumiy)).replace(",", " "),
    }

@login_required(login_url='login')
def HomePage(request):
    logger.info("HomePage view is running")
    today = timezone.now().date()
    one_week_ago = today - timedelta(days=7)
    one_month_ago = today - timedelta(days=30)

    try:
        # Daily expenses for the last 7 days
        daily_expenses = Expense.objects.filter(
            created__date__gte=one_week_ago
        ).annotate(
            date=TruncDate('created')
        ).values('date').annotate(
            total=Sum('amount')
        ).order_by('date')

        # Monthly expenses by type
        expense_by_type = Expense.objects.filter(
            created__date__gte=one_month_ago
        ).values(
            'expense_type__name'
        ).annotate(
            total=Sum('amount')
        ).order_by('-total')

        # Building expenses
        building_expenses = Expense.objects.filter(
            created__date__gte=one_month_ago,
            building__isnull=False
        ).values(
            'building__name'
        ).annotate(
            total=Sum('amount')
        ).order_by('-total')

        # Format expense data for charts
        expense_data = {
            'daily': {
                'dates': [expense['date'].strftime('%Y-%m-%d') for expense in daily_expenses],
                'amounts': [float(expense['total']) for expense in daily_expenses]
            },
            'by_type': {
                    'types': [expense['expense_type__name'] or 'Nomalum' for expense in expense_by_type],
                'amounts': [float(expense['total']) for expense in expense_by_type]
            },
            'by_building': {
                    'buildings': [expense['building__name'] or 'Nomalum' for expense in building_expenses],
                'amounts': [float(expense['total']) for expense in building_expenses]
            }
        }
    except Exception as e:
        logger.info(f"Error processing expense data: {e}")
        expense_data = {
            'daily': {'dates': [], 'amounts': []},
            'by_type': {'types': [], 'amounts': []},
            'by_building': {'buildings': [], 'amounts': []}
        }

    # Get other data for the template
    # Dinamik eshitilgan manbalar
    heard_qs = ClientInformation.objects.values('heard').annotate(total=Count('id')).order_by('-total')
    heard_labels = []
    heard_counts = []
    for item in heard_qs:
        label = item['heard'] or "Noma'lum"
        heard_labels.append(label)
        heard_counts.append(item['total'])

    active_contracts = ConsultingContract.objects.exclude(
        status__in=[ConsultingContract.StatusChoices.COMPLETED, ConsultingContract.StatusChoices.CANCELLED]
    ).count()
    completed_contracts = ConsultingContract.objects.filter(status=ConsultingContract.StatusChoices.COMPLETED).count()

    context = {
        'tushum': tushum_view(request),
        'client_count': ClientInformation.objects.count(),
        'building_count': Building.objects.count(),
        'month_client': ClientInformation.objects.filter(created__month=timezone.now().month).count(),
        'contract': active_contracts,
        'contract_f': completed_contracts,
        # Qarzdorlar - yangi ConsultingContract modelidan
        'qarz': sum([contract.remaining_amount for contract in ConsultingContract.objects.all() if contract.remaining_amount > 0]),
        'debtors': {
            'debtor': ConsultingContract.objects.filter(amount_paid__lt=F('total_service_fee')).count(),
            'nodebtor': ConsultingContract.objects.filter(amount_paid__gte=F('total_service_fee')).count()
        },
        'heard_labels': json.dumps(heard_labels, ensure_ascii=False),
        'heard_counts': json.dumps(heard_counts),
        'week_list': [d.strftime('%Y-%m-%d') for d in [(today - timedelta(days=x)) for x in range(6, -1, -1)]],
        'week_client': [ClientInformation.objects.filter(created__date=d).count() for d in [(today - timedelta(days=x)) for x in range(6, -1, -1)]],
        'expense_data': json.dumps(expense_data)
    }

    return render(request, 'index.html', context)

def build_filter_params(request):
    """Build filter parameters from request"""
    filters = {}
    params_dict = {}
    
    building_id = request.GET.get("building")
    city_id = request.GET.get("city")
    status = request.GET.get("status")
    
    if city_id and city_id.isdigit():
        filters["building__city__id"] = city_id
        params_dict["city"] = city_id
        
    if building_id and building_id.isdigit():
        filters["building__id"] = building_id
        params_dict["building"] = building_id
        
    if status:
        filters['home__busy'] = True if status == 'occupied' else False
        params_dict["status"] = status
        
    

    
    return filters, params_dict

def build_filter_url(params_dict):
    """Build filter URL from parameters dictionary"""
    if not params_dict:
        return ""
    
    url_parts = []
    for key, value in params_dict.items():
        url_parts.append(f"{key}={value}")
    
    return "?" + "&".join(url_parts)


def normalize_phone(phon):
    if not phon:
        return None
    # Float ko'rinishida kelsa, butun qismini olish
    if isinstance(phon, float):
        phon = str(int(phon))
    phone = str(phon)
    digits = re.sub(r'\D', '', phone)
    # Agar raqam 9 tadan ko'p bo'lsa, oxirgi 9 tasini olish
    if len(digits) > 9:
        digits = digits[-9:]
    if len(digits) == 9:
        return '+998' + digits
    elif len(digits) == 12 and digits.startswith('998'):
        return '+' + digits
    else:
        return None
    

async def send_sms(phone, sms: str):
    if not phone:
        return False
        
    try:
        url = "https://notify.eskiz.uz/api/message/sms/send"
        params = {
            "mobile_phone": f"{phone}",
            "message": f"{sms}",
            "from": "4546",
        }
        headers = {
            "Authorization": "Bearer eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJleHAiOjE3NTI3NDQwMzQsImlhdCI6MTc1MDE1MjAzNCwicm9sZSI6InVzZXIiLCJzaWduIjoiZGI1NjYwNWVjMTQ2ZjMzNzVjOWYyMDA0N2ZlNzY1M2MxYzlkMzQ4NTZjM2FmODBlZjljYWI3MGE5NDE3YmJkZCIsInN1YiI6IjExMzgxIn0.EeOcYIq0POwawE1SYQ7nhapx5VdFgZqGGsarOFtVUJw",
        }

        async with aiohttp.ClientSession() as session:
            async with session.post(
                url, data=params, headers=headers, ssl=False
            ) as response:
                if response.status == 200:
                    data = await response.json()
                    logger.info("SMS yuborildi, javob:", data)
                    return True
                else:
                    logger.info(
                        f"SMS yuborishda xatolik:\nstatus: {response.status}\njavob: {await response.text()}"
                    )
                    return False
    except Exception as e:
        logger.info("Xatolik:", e)
        return False

@login_required(login_url='login')
def ClientPage(request):
    clients = ClientInformation.objects.all().order_by("-created")
    search_value = ""
    
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "search":
            search_term = request.POST.get("search", "").strip()
            if search_term:
                search_value = search_term
                clients = clients.filter(
                    Q(full_name__icontains=search_term) |
                    Q(phone__icontains=search_term)
                )
        
        elif action == "sms":
            # SMS yuborish funksiyasi
            recipient_type = request.POST.get("recipient_type")
            sms_text = request.POST.get("sms-text")
            custom_recipients = request.POST.get("custom_recipients")
            
            if sms_text:
                phone_numbers = []
                
                if recipient_type == "all":
                    phone_numbers = [client.phone for client in ClientInformation.objects.all()]
                elif recipient_type == "telegram":
                    phone_numbers = [client.phone for client in ClientInformation.objects.filter(heard="Telegramda")]
                elif recipient_type == "instagram":
                    phone_numbers = [client.phone for client in ClientInformation.objects.filter(heard="Instagramda")]
                elif recipient_type == "youtube":
                    phone_numbers = [client.phone for client in ClientInformation.objects.filter(heard="YouTubeda")]
                elif recipient_type == "people":
                    phone_numbers = [client.phone for client in ClientInformation.objects.filter(heard="Odamlar orasida")]
                elif recipient_type == "custom" and custom_recipients:
                    phone_numbers = [phone.strip() for phone in custom_recipients.split(",")]
                
                if phone_numbers:
                    # SMS yuborish logikasi bu yerda bo'ladi
                    # Hozircha faqat success message ko'rsatamiz
                    messages.success(request, f"{len(phone_numbers)} ta mijozga SMS yuborildi.")
                else:
                    messages.warning(request, "SMS yuborish uchun telefon raqamlar topilmadi.")
            else:
                messages.warning(request, "SMS matni kiritilmadi.")
        
        elif action == "sms-one":
            # Bitta mijozga SMS yuborish
            custom_recipients = request.POST.get("custom_recipients")
            sms_text = request.POST.get("sms-text")
            
            if sms_text and custom_recipients:
                # SMS yuborish logikasi bu yerda bo'ladi
                messages.success(request, f"{custom_recipients} raqamiga SMS yuborildi.")
            else:
                messages.warning(request, "SMS matni yoki telefon raqam kiritilmadi.")

    # GET parametrlari orqali filterlash
    filter_param = (request.GET.get("filter") or "").strip()
    date_param = (request.GET.get("date") or "").strip()
    phone_param = (request.GET.get("phone") or "").strip()
    search_param = (request.GET.get("search") or request.GET.get("q") or "").strip()
    search_value = search_param

    # Unique heard choices (trimmed) for filters/selects
    heard_raw = ClientInformation.objects.values_list("heard", flat=True).distinct()
    heard_choices = sorted({(h or "").strip() or "Noma'lum" for h in heard_raw})
    
    if filter_param:
        heard_options = {
            "0": "Telegramda",
            "1": "Instagramda", 
            "2": "YouTubeda",
            "3": "Odamlar orasida",
            "4": "Xech qayerda"
        }
        heard_value = heard_options.get(filter_param, filter_param)
        clients = clients.filter(heard=heard_value)
    
    if date_param:
        try:
            filter_date = datetime.strptime(date_param, "%Y-%m-%d").date()
            clients = clients.filter(created__date=filter_date)
        except ValueError:
            pass
    
    if phone_param:
        cleaned_phone = re.sub(r"\D", "", phone_param)
        if cleaned_phone:
            clients = clients.filter(
                Q(phone__icontains=phone_param) |
                Q(phone__icontains=cleaned_phone)
            )
    
    if search_param:
        clients = clients.filter(
            Q(full_name__icontains=search_param) |
            Q(phone__icontains=search_param)
        )

    # Yangi → eski tartibda (yangi yaratilgan birinchi bo'lib ko'rinadi)
    clients_ordered = clients.order_by("-created")
    
    # Paginatsiya - har sahifada 15 ta mijoz
    paginator = Paginator(clients_ordered, 15)
    page = request.GET.get('page', 1)
    
    try:
        clients = paginator.page(page)
    except PageNotAnInteger:
        clients = paginator.page(1)
    except EmptyPage:
        clients = paginator.page(paginator.num_pages)

    # Filter parametrlarini saqlash
    filters, params_dict = build_filter_params(request)
    filter_params = build_filter_url(params_dict)
    
    # Statistikalar (barcha mijozlar bo'yicha, paginatsiyasiz)
    today = timezone.now().date()
    all_clients = ClientInformation.objects.all()
    active_clients = all_clients.filter(created__date__gte=today - timedelta(days=30))
    today_clients = all_clients.filter(created__date=today).count()

    context = {
        "client": clients,
        "search_value": search_value,
        "active_clients": active_clients,
        "today_clients": today_clients,
        "filter_params": filter_params,
        "selected_filter": filter_param,
        "selected_date": date_param,
        "selected_phone": phone_param,
        "heard_choices": heard_choices,
    }

    return render(request, "client/client.html", context)


@login_required(login_url='login')
@ceoadmin_required
def ClientCreate(request):
        
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        middle_name = request.POST.get("middle_name", "").strip() or None
        phone = request.POST.get("phone")
        phone2 = request.POST.get("phone2", "")
        heard = request.POST.get("heard")

        if not first_name:
            messages.warning(request, "Ism kiritilmadi.")
            return render(request=request, template_name="client/create.html")
        
        if not last_name:
            messages.warning(request, "Familiya kiritilmadi.")
            return render(request=request, template_name="client/create.html")
            
        if not phone:
            messages.warning(request, "Telefon raqami kiritilmadi.")
            return render(request=request, template_name="client/create.html")
            
        if not heard:
            messages.warning(request, "Qayerda eshitgani kiritilmadi.")
            return render(request=request, template_name="client/create.html")

        # Normalize phone numbers
        phone_clean = normalize_phone(phone)
        phone2_clean = normalize_phone(phone2) if phone2 else None
        
        if not phone_clean:
            messages.warning(request, "Telefon raqami noto'g'ri formatda.")
            return render(request=request, template_name="client/create.html")

        # Build full_name for checking
        name_parts = [last_name, first_name]
        if middle_name:
            name_parts.append(middle_name)
        full_name = ' '.join(name_parts)

        # Check if client already exists
        existing_client = ClientInformation.objects.filter(
            Q(phone=phone_clean) | Q(first_name=first_name, last_name=last_name)
        ).first()
        
        if existing_client:
            messages.warning(request, "Bu mijoz allaqachon mavjud.")
            return render(
                request=request, 
                template_name="client/create.html", 
                context={"existing_client": existing_client}
            )

        # Get additional fields
        passport_number = request.POST.get("passport_number", "").strip() or None
        passport_issue_date = request.POST.get("passport_issue_date", "").strip() or None
        passport_expiry_date = request.POST.get("passport_expiry_date", "").strip() or None
        passport_issue_place = request.POST.get("passport_issue_place", "").strip() or None
        birth_date = request.POST.get("birth_date", "").strip() or None
        address = request.POST.get("address", "").strip() or None

        # Create new client
        ClientInformation.objects.create(
            first_name=first_name,
            last_name=last_name,
            middle_name=middle_name,
            phone=phone_clean, 
            phone2=phone2_clean,
            passport_number=passport_number,
            passport_issue_date=passport_issue_date,
            passport_expiry_date=passport_expiry_date,
            passport_issue_place=passport_issue_place,
            birth_date=birth_date,
            address=address,
            heard=heard
        )
        messages.success(request, "Mijoz muvaffaqiyatli yaratildi.")
        return redirect("client")

    return render(request=request, template_name="client/create.html")

@login_required(login_url='login')
@ceoadmin_required
def ClientDelete(request, id):
    try:
        client_instance = get_object_or_404(ClientInformation, pk=id)
        
        # Check if client has contracts (both old Client and new ConsultingContract)
        has_old_contracts = Client.objects.filter(client=client_instance).exists()
        has_new_contracts = ConsultingContract.objects.filter(client=client_instance).exists()
        
        if has_old_contracts or has_new_contracts:
            return JsonResponse({
                "ok": False,
                "message": "Mijozni olib tashlash mumkin emas. Sababi bu mijoz nomiga shartnoma rasmiylashtirilgan.",
            })
        
        client_instance.delete()
        messages.success(request, "Mijoz muvaffaqiyatli o'chirildi.")
        return JsonResponse({"ok": True, "message": "Mijoz muvaffaqiyatli o'chirildi."})
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Client delete error: {str(e)}", exc_info=True)
        return JsonResponse({
            "ok": False,
            "message": f"O'chirishda xatolik yuz berdi: {str(e)}",
        })

@login_required(login_url='login')
@ceoadmin_required
def ClientEdit(request, id):
    client = get_object_or_404(ClientInformation, pk=id)
    
    # Filter parametrlarini saqlash
    filters, params_dict = build_filter_params(request)
    filter_params = build_filter_url(params_dict)
    next_url = request.GET.get('next')
    
    if request.method == "POST":
        first_name = request.POST.get("first_name", "").strip()
        last_name = request.POST.get("last_name", "").strip()
        middle_name = request.POST.get("middle_name", "").strip() or None
        phone = request.POST.get("phone")
        phone2 = request.POST.get("phone2", "").strip() or None
        heard = request.POST.get("heard")
        passport_number = request.POST.get("passport_number", "").strip() or None
        passport_issue_date = request.POST.get("passport_issue_date", "").strip() or None
        passport_expiry_date = request.POST.get("passport_expiry_date", "").strip() or None
        passport_issue_place = request.POST.get("passport_issue_place", "").strip() or None
        birth_date = request.POST.get("birth_date", "").strip() or None
        address = request.POST.get("address", "").strip() or None
        
        if not first_name:
            messages.warning(request, "Ism kiritilmadi.")
        elif not last_name:
            messages.warning(request, "Familiya kiritilmadi.")
        elif not phone:
            messages.warning(request, "Telefon raqam kiritilmadi.")
        elif not heard:
            messages.warning(request, "Qayerda eshitgani tanlanmadi.")
        else:
            # Telefon raqam formatini tekshirish
            phone_clean = normalize_phone(phone)
            phone2_clean = normalize_phone(phone2) if phone2 else None
            
            if not phone_clean:
                messages.warning(request, "Telefon raqam noto'g'ri formatda.")
            else:
                # Boshqa mijozda bu telefon raqam mavjudligini tekshirish
                existing_client = ClientInformation.objects.filter(phone=phone_clean).exclude(pk=id).first()
                if existing_client:
                    messages.warning(request, "Bu telefon raqam bilan boshqa mijoz mavjud.")
                else:
                    # Mijozni yangilash
                    client.first_name = first_name
                    client.last_name = last_name
                    client.middle_name = middle_name
                    client.phone = phone_clean
                    client.phone2 = phone2_clean
                    client.passport_number = passport_number
                    client.passport_issue_date = passport_issue_date
                    client.passport_expiry_date = passport_expiry_date
                    client.passport_issue_place = passport_issue_place
                    # birth_date ni har doim yangilash (bo'sh bo'lsa ham None sifatida saqlanadi)
                    client.birth_date = birth_date.strip() if birth_date and birth_date.strip() else None
                    client.address = address
                    client.heard = heard
                    client.save()
                    
                    messages.success(request, "Mijoz ma'lumotlari muvaffaqiyatli yangilandi.")
                    
                    # next parametri bo'yicha qaytish
                    if next_url:
                        return redirect(next_url)
                    else:
                        return redirect('client')
    
    context = {
        "client": client,
        "filter_params": build_filter_url(filter_params),
        "next_url": next_url,
    }
    
    return render(request, "client/edit.html", context)



def _clean_currency_value(value, allow_empty=True):
    if value is None:
        return 0
    text = str(value).strip()
    if not text:
        return 0 if allow_empty else None
    digits = re.sub(r"[^\d]", "", text)
    if not digits:
        return 0
    return int(digits)


def _parse_contract_date(value):
    try:
        if not value:
            return timezone.now().date()
        cleaned = value.replace(" ", "")
        if "." in cleaned:
            return datetime.strptime(cleaned, "%d.%m.%Y").date()
        return datetime.strptime(cleaned, "%Y-%m-%d").date()
    except ValueError:
        raise ValueError("Shartnoma sanasi noto'g'ri formatda. (DD.MM.YYYY)")


def _parse_int_safe(val, default=0):
    try:
        if val is None:
            return default
        return int(str(val).replace(" ", "").replace(",", ""))
    except (TypeError, ValueError):
        return default


def _handle_uploaded_images(request, prefix, max_images, contract_number=None, client_name=None):
    """Rasmlarni yuklab, saqlab, JSON formatida qaytaradi"""
    images = []
    # Filenamelar uchun prefiks: faqat mijoz ismi, kontrakt raqami va uuid ishlatmaymiz
    base_prefix = slugify(client_name) if client_name else prefix

    for i in range(max_images):
        file_key = f"{prefix}_{i}"
        if file_key in request.FILES:
            uploaded_file = request.FILES[file_key]
            # Fayl nomini yaratish (faqat mijoz nomi + kengaytma)
            file_ext = os.path.splitext(uploaded_file.name)[1]
            file_name = f"{prefix}/{base_prefix}{file_ext}"
            # Faylni saqlash
            file_path = default_storage.save(file_name, ContentFile(uploaded_file.read()))
            # Media URL bilan to'liq yo'l yaratish
            # file_path: passport_image/xxx.png (default_storage.save() faqat relative path qaytaradi)
            # MEDIA_URL: media/
            # Natija: /media/passport_image/xxx.png
            media_url = settings.MEDIA_URL.rstrip('/')
            # file_path dan media/ prefiksini olib tashlash (agar bor bo'lsa)
            # default_storage.save() faqat relative path qaytaradi: passport_image/xxx.png
            if file_path.startswith('media/'):
                file_path = file_path[6:]  # 'media/' ni olib tashlash
            elif file_path.startswith('/media/'):
                file_path = file_path[7:]  # '/media/' ni olib tashlash
            
            # Agar file_path allaqachon /media/ bilan boshlanmasa, qo'shamiz
            if not file_path.startswith('/media/'):
                full_path = f"/{media_url}/{file_path}" if media_url else f"/{file_path}"
            else:
                full_path = file_path
            images.append(full_path)
    return images

def _extract_consulting_contract_payload(form, min_contract_number=None, existing_id=None, auto_generate=True):
    errors = []

    # Shartnoma raqami va sanasi avtomatik yaratiladi (create holatida)
    contract_number = None
    contract_date = None
    
    if auto_generate and not existing_id:
        # Avtomatik yaratish (create holatida)
        contract_number = min_contract_number
        contract_date = timezone.now().date()
    else:
        # Edit holatida mavjud qiymatlarni saqlash
        if existing_id:
            try:
                existing_contract = ConsultingContract.objects.get(pk=existing_id)
                contract_number = existing_contract.contract_number
                contract_date = existing_contract.contract_date
            except ConsultingContract.DoesNotExist:
                pass

    contract_location = (form.get("contract_location") or "Xiva").strip()
    client_first_name = (form.get("client_first_name") or "").strip()
    client_last_name = (form.get("client_last_name") or "").strip()
    client_middle_name = (form.get("client_middle_name") or "").strip() or None
    passport_number = (form.get("passport_number") or "").strip()
    passport_issue_date = (form.get("passport_issue_date") or "").strip() or None
    passport_expiry_date = (form.get("passport_expiry_date") or "").strip() or None
    passport_issue_place = (form.get("passport_issue_place") or "").strip() or None
    birth_date = (form.get("birth_date") or "").strip() or None
    client_address = (form.get("client_address") or "").strip()
    visa_type = (form.get("visa_type") or "Ishchi viza").strip()
    service_name = (form.get("service_name") or "Angliya ishchi viza paketi").strip()

    # Build full_name for backward compatibility
    name_parts = [client_last_name, client_first_name]
    if client_middle_name:
        name_parts.append(client_middle_name)
    client_full_name = ' '.join(name_parts)

    if not client_first_name:
        errors.append("Mijoz ismi kiritilishi shart.")
    if not client_last_name:
        errors.append("Mijoz familiyasi kiritilishi shart.")
    if not passport_number:
        errors.append("Passport raqami kiritilishi shart.")
    else:
        # AA1234567 formatini majburiy qilish (2 harf, 7 raqam)
        if not re.match(r"^[A-Za-z]{2}[0-9]{1,7}$", passport_number):
            errors.append("Passport raqami 'AA1234567' ko'rinishida bo'lishi kerak (avval 2 harf, so'ng raqamlar).")
    if not visa_type:
        errors.append("Visa yo'nalishi kiritilishi shart.")
    if not service_name:
        errors.append("Xizmat nomi kiritilishi shart.")

    heard = (form.get("heard") or "").strip()

    def _clean_phone(raw):
        if not raw:
            return None
        val = str(raw).strip()
        # Agar faqat kod bo'lsa yoki juda qisqa bo'lsa, bo'sh deb qaraymiz
        if val in {"+998", "998", "+998 ", "998 "}:
            return None
        normalized = normalize_phone(val)
        return normalized

    phone_primary = _clean_phone(form.get("phone_primary"))
    phone_secondary = _clean_phone(form.get("phone_secondary"))

    if not phone_primary:
        errors.append("Asosiy telefon raqami noto'g'ri kiritildi.")
    if not heard:
        errors.append("Qayerda eshitgani kiritilishi shart.")

    total_fee = _clean_currency_value(form.get("total_service_fee"))
    initial_payment = _clean_currency_value(form.get("initial_payment_amount"))
    second_payment = _clean_currency_value(form.get("post_interview_payment_amount"))
    refund_amount = _clean_currency_value(form.get("refund_amount"))
    amount_paid = _clean_currency_value(form.get("amount_paid"))

    if second_payment == 0 and total_fee and initial_payment is not None:
        second_payment = max(total_fee - initial_payment, 0)

    try:
        initial_due_days = int(form.get("initial_payment_due_days") or 3)
        second_due_days = int(form.get("post_interview_due_days") or 3)
        duration_months = int(form.get("service_duration_months") or 8)
    except ValueError:
        initial_due_days = 3
        second_due_days = 3
        duration_months = 8
        errors.append("Kun va oy qiymatlari butun son bo'lishi kerak.")

    status = form.get("status") or ConsultingContract.StatusChoices.PREPARATION
    if status not in dict(ConsultingContract.StatusChoices.choices):
        status = ConsultingContract.StatusChoices.PREPARATION

    # Email va password
    email = (form.get("email") or "").strip() or None
    password = (form.get("password") or "").strip() or None
    
    # Mijoz ma'lumotlari (ClientInformation uchun)
    client_data = {
        "first_name": client_first_name,
        "last_name": client_last_name,
        "middle_name": client_middle_name,
        "phone": phone_primary,
        "phone2": phone_secondary,
        "passport_number": passport_number,
        "passport_issue_date": passport_issue_date,
        "passport_issue_place": passport_issue_place,
        "passport_expiry_date": passport_expiry_date,
        "birth_date": birth_date.strip() if birth_date and birth_date.strip() else None,  # Bo'sh bo'lsa ham None sifatida qo'shamiz
        "address": client_address,
        "email": email,
        "password": password,
        "heard": heard,
    }

    # Shartnoma ma'lumotlari
    payload = {
        "contract_number": contract_number,
        "contract_date": contract_date,
        "contract_location": contract_location,
        # Yangi maydonlar
        "client_first_name": client_first_name,
        "client_last_name": client_last_name,
        "client_middle_name": client_middle_name,
        # Eski maydonlar (backward compatibility)
        "client_full_name": client_full_name,
        "passport_number": passport_number,
        "passport_issue_date": passport_issue_date,
        "passport_issue_place": passport_issue_place,
        "passport_expiry_date": passport_expiry_date,
        "client_address": client_address,
        "phone_primary": phone_primary,
        "phone_secondary": phone_secondary,
        "service_name": service_name,
        "service_country": form.get("service_country") or "Angliya",
        "visa_type": visa_type,
        "service_description": form.get("service_description") or None,
        "total_service_fee": total_fee,
        "initial_payment_amount": initial_payment,
        "initial_payment_due_days": initial_due_days,
        "post_interview_payment_amount": second_payment,
        "post_interview_due_days": second_due_days,
        "refund_amount": refund_amount,
        "service_duration_months": duration_months,
        "amount_paid": amount_paid,
        "status": status,
        "notes": form.get("notes") or None,
    }

    return payload, errors, client_data


@login_required(login_url='login')
def ContractPage(request):

    contract_queryset = ConsultingContract.objects.all()
    status_filter = request.GET.get("status") or ""
    search_value = request.GET.get("q", "").strip()
    created_by_filter = request.GET.get("created_by") or ""

    if status_filter in dict(ConsultingContract.StatusChoices.choices):
        contract_queryset = contract_queryset.filter(status=status_filter)

    if search_value:
        contract_queryset = contract_queryset.filter(
            Q(client_full_name__icontains=search_value) |
            Q(passport_number__icontains=search_value) |
            Q(phone_primary__icontains=search_value) |
            Q(contract_number__icontains=search_value) |
            Q(service_name__icontains=search_value) |
            Q(client__full_name__icontains=search_value) |
            Q(client__phone__icontains=search_value) |
            Q(client__passport_number__icontains=search_value)
        )
    if created_by_filter and created_by_filter.isdigit():
        contract_queryset = contract_queryset.filter(created_by_id=int(created_by_filter))

    # Yangi → eski tartibda (yangi yaratilgan birinchi bo'lib ko'rinadi)
    contracts_ordered = contract_queryset.order_by("-created_at")
    
    # Paginatsiya - har sahifada 20 ta shartnoma
    paginator = Paginator(contracts_ordered, 15)
    page = request.GET.get('page', 1)
    
    try:
        contracts = paginator.page(page)
    except PageNotAnInteger:
        contracts = paginator.page(1)
    except EmptyPage:
        contracts = paginator.page(paginator.num_pages)
    
    aggregates = contract_queryset.aggregate(
        total_fee=Sum("total_service_fee"),
        total_paid=Sum("amount_paid")
    )
    total_contracts = ConsultingContract.objects.count()
    active_contracts = ConsultingContract.objects.exclude(
        status__in=[ConsultingContract.StatusChoices.COMPLETED, ConsultingContract.StatusChoices.CANCELLED]
    ).count()
    completed_contracts = ConsultingContract.objects.filter(
        status=ConsultingContract.StatusChoices.COMPLETED
    ).count()
    users_with_contracts = User.objects.filter(created_contracts__isnull=False).distinct()

    total_fee = aggregates.get("total_fee") or 0
    total_paid = aggregates.get("total_paid") or 0
    outstanding = total_fee - total_paid if total_fee > total_paid else 0

    context = {
        "contracts": contracts,
        "search_value": search_value,
        "status_filter": status_filter,
        "created_by_filter": created_by_filter,
        "status_choices": ConsultingContract.StatusChoices.choices,
        "users_with_contracts": users_with_contracts,
        "total_contracts": total_contracts,
        "active_contracts": active_contracts,
        "completed_contracts": completed_contracts,
        "total_fee": total_fee,
        "outstanding": outstanding,
    }

    return render(request, "contract/contract.html", context)

@login_required(login_url='login')
def ContractDetailsAPI(request, id):
    """Shartnoma tafsilotlarini JSON formatida qaytaradi"""
    contract = get_object_or_404(ConsultingContract, pk=id)
    client_info = contract.client
    
    # Mijoz ma'lumotlari
    client_data = {
        "first_name": client_info.first_name if client_info else contract.client_first_name or "",
        "last_name": client_info.last_name if client_info else contract.client_last_name or "",
        "middle_name": client_info.middle_name if client_info else contract.client_middle_name or "",
        "phone": client_info.phone if client_info else contract.phone_primary or "",
        "phone2": client_info.phone2 if client_info else contract.phone_secondary or "",
        "email": client_info.email if client_info else "",
        "password": client_info.password if client_info else "",
        "passport_number": client_info.passport_number if client_info else contract.passport_number or "",
        "passport_issue_date": client_info.passport_issue_date if client_info else getattr(contract, "passport_issue_date", "") or "",
        "passport_expiry_date": getattr(contract, "passport_expiry_date", "") or "",
        "passport_issue_place": client_info.passport_issue_place if client_info else contract.passport_issue_place or "",
        "birth_date": client_info.birth_date if client_info and client_info.birth_date else "",
        "address": client_info.address if client_info else contract.client_address or "",
        "heard": client_info.heard if client_info else "",
    }
    
    # Family members
    family_members = []
    for member in contract.family_members.all():
        family_members.append({
            "first_name": member.first_name,
            "last_name": member.last_name,
            "middle_name": member.middle_name or "",
            "relationship": member.get_relationship_display(),
            "passport_number": member.passport_number or "",
            "passport_issue_date": member.passport_issue_date or "",
        "passport_expiry_date": getattr(member, "passport_expiry_date", "") or "",
            "passport_issue_place": member.passport_issue_place or "",
            "birth_date": member.birth_date if member.birth_date else "",
            "phone": member.phone or "",
            "notes": member.notes or "",
        })
    
    data = {
        "number": contract.contract_number,
        "date": contract.contract_date.strftime("%d.%m.%Y"),
        "location": contract.contract_location,
        "status": contract.get_status_display(),
        "client": client_data,
        "service": {
            "name": contract.service_name,
            "country": contract.service_country,
            "visa_type": contract.visa_type,
            "description": contract.service_description or "",
        },
        "payment": {
            "total": contract.total_service_fee,
            "paid": contract.amount_paid,
            "remaining": contract.remaining_amount,
            "initial": contract.initial_payment_amount,
            "post_interview": contract.post_interview_payment_amount,
        },
        "passport_images": [
            img if (img.startswith('/media/') or img.startswith('http')) 
            else (f"/media/{img.replace('/media/', '').replace('media/', '').lstrip('/')}" if img else "")
            for img in (contract.passport_images or []) if img
        ],
        "visa_images": [
            img if (img.startswith('/media/') or img.startswith('http')) 
            else (f"/media/{img.replace('/media/', '').replace('media/', '').lstrip('/')}" if img else "")
            for img in (contract.visa_images or []) if img
        ],
        "completed_contract_images": [
            img if (img.startswith('/media/') or img.startswith('http')) 
            else (f"/media/{img.replace('/media/', '').replace('media/', '').lstrip('/')}" if img else "")
            for img in (contract.completed_contract_images or []) if img
        ],
        "family_members": family_members,
        "notes": contract.notes or "",
    }
    
    return JsonResponse(data)

# @login_required(login_url='login')
# def JadvalPage(request, id):
#     """
#     To'lovlar jadvalini ko'rsatish uchun view funksiya.
#     Bu funksiya shartnoma bo'yicha to'lovlar jadvalini ko'rsatadi va
#     frontend uchun zarur bo'lgan barcha ma'lumotlarni tayyorlaydi.
#     """
#     if request.user.username == "financeadmin":
#         return redirect("login")
    
#     contract = get_object_or_404(Client, pk=id)
#     rasrochka = Rasrochka.objects.filter(client=contract).order_by("month")
#     home_price = contract.home_price
    
#     # # Check if user is ceoadmin for special permissions
#     # is_ceo_admin = request.user.username == "ceoadmin"
    
#     # Get next URL for back navigation with filters
#     next_url = get_contract_next_url(request)
#     filters, params_dict = build_contract_filter_params(request)
#     filter_params = build_contract_filter_url(params_dict)
    
#     if request.method == "POST":
#         payment_type = request.POST.get("payment-type")
#         if payment_type == "monthly":
#             return handle_monthly_payment(request, contract, id, next_url)
#         elif payment_type == "custom":
#             return handle_custom_payment(request, contract, id, next_url)
#         else:
#             messages.warning(request, "To'lov turi aniqlanmadi")
    
#     payment_schedule_data = []
#     total_amount = 0
#     total_paid = 0
#     total_remaining = 0
    
#     for payment in rasrochka:
#         payment_data = {
#             'id': payment.id,
#             'month': payment.month,
#             'amount': int(str(payment.amount).replace(' ', '')),
#             'amount_paid': int(str(payment.amount_paid).replace(' ', '')),
#             'qoldiq': int(str(payment.qoldiq).replace(' ', '')),
#             'date': payment.date.strftime('%Y-%m-%d'),
#             'pay_date': payment.pay_date.strftime('%Y-%m-%d') if payment.pay_date else None,
#             'is_initial': payment.month == 0,
#             'is_paid': payment.qoldiq == 0,
#             'is_partially_paid': payment.amount_paid > 0 and payment.qoldiq > 0,
#             'can_pay': payment.qoldiq > 0 and contract.status != "Tugallangan",
#             # 'can_admin_edit': is_ceo_admin and payment.amount_paid > 0  # Only CEO can edit paid amounts
#         }
#         payment_schedule_data.append(payment_data)
#         total_amount += payment_data['amount']
#         total_paid += payment_data['amount_paid']
#         total_remaining += payment_data['qoldiq']
    
#     months_count = len([p for p in payment_schedule_data if not p['is_initial']])
    
#     contract_data = {
#         'id': contract.id,
#         'contract_number': contract.contract,
#         'status': contract.status,
#         'pay_date': contract.pay_date,
#         'total_price': int(str(home_price).replace(' ', '')),
#         'client_name': contract.client.full_name if contract.client else '',
#         'home_info': f"{contract.home.building.name} - {contract.home.home.home_number}" if contract.home else ''
#     }
    
#     context = {
#         "contract": contract,
#         "rasrochka": rasrochka,
#         "price": int(str(home_price).replace(' ', '')),
#         "qolgan_price": int(str(contract.residual).replace(' ', '')),
#         "pk": id,
#         "payment_schedule_data": json.dumps(payment_schedule_data),
#         "total_amount": total_amount,
#         "total_paid": total_paid,
#         "total_remaining": int(str(contract.residual).replace(' ', '')),
#         "months_count": months_count,
#         "contract_data": json.dumps(contract_data),
#         "next_url": next_url,
#         "filter_params": filter_params,
#         # "is_ceo_admin": is_ceo_admin  # Pass to template
#     }
    
#     return render(request, "contract/list.html", context)

# def handle_monthly_payment(request, contract: Client, contract_id, next_url):
#     """
#     Oylik to'lovni qayta ishlash uchun funksiya.
#     Bu funksiya ma'lum bir oy uchun to'lovni qabul qiladi.
#     """
#     # To'lov ma'lumotlarini olish
#     debt_id = request.POST.get("debt-id")
#     summa = request.POST.get("amount")
    
#     # Ma'lumotlar to'liqligini tekshirish
#     if not debt_id or not summa:
#         messages.warning(request, "To'lov ma'lumotlari to'liq emas")
#         return render(
#             request,
#             "contract/list.html",
#             {
#                 "contract": contract,
#                 "rasrochka": Rasrochka.objects.filter(client=contract).order_by("date"),
#                 "price": contract.home.home.price * contract.home.home.field,
#                 "pk": contract_id
#             },
#         )
    
#     # To'lovni topish
#     rasrorchka_obj = Rasrochka.objects.filter(pk=debt_id)
    
#     if rasrorchka_obj.exists():
#         value = rasrorchka_obj.first()
        
#         if value.client:
#             # To'lovni qo'shish - Decimal conversion qo'shildi
#             summa_decimal = Decimal(str(int(summa.replace(' ', '').replace(',', ''))))
#             value.amount_paid += summa_decimal
#             value.pay_date = datetime.now()
            
#             # Qoldiqni yangilash
#             value.qoldiq = max(0, value.amount - value.amount_paid)
#             value.save()
            
#             # SMS matni tayyorlash
#             # formatted_summa = "{:,}".format(int(summa.replace(' ', '').replace(",", " ")))
            
#             # Shartnoma qoldig'ini yangilash - Decimal conversion qo'shildi
#             contract.residual -= summa_decimal
            
#             # Agar qoldiq 0 bo'lsa, shartnomani tugallangan deb belgilash
#             if contract.residual <= 0:
#                 contract.debt = False
#                 if contract.status == 'Rasmiylashtirilgan':
#                     contract.status = 'Tugallangan'
#             else:
#                 contract.debt = True
                
#             contract.save()
            
#             formatted_residual = "{:,}".format(int(contract.residual)).replace(",", " ")
            
#             messages.success(request, "To'lov muvaffaqiyatli qabul qilindi")
            
#             # Frontend uchun to'lovlar jadvalini tayyorlash
#             payment_schedule_data = []
#             total_amount = 0
#             total_paid = 0
#             total_remaining = 0
            
#             rasrochka = Rasrochka.objects.filter(client=contract).order_by("month")
            
#             for payment in rasrochka:
#                 payment_data = {
#                     'id': payment.id,
#                     'month': payment.month,
#                     'amount': float(payment.amount),
#                     'amount_paid': float(payment.amount_paid),
#                     'qoldiq': float(payment.qoldiq),
#                     'date': payment.date.strftime('%Y-%m-%d'),
#                     'pay_date': payment.pay_date.strftime('%Y-%m-%d') if payment.pay_date else None,
#                     'is_initial': payment.month == 0,
#                     'is_paid': payment.qoldiq == 0,
#                     'can_pay': payment.qoldiq > 0 and contract.status != "Tugallangan"
#                 }
#                 payment_schedule_data.append(payment_data)
                
#                 total_amount += float(payment.amount)
#                 total_paid += float(payment.amount_paid)
#                 total_remaining += float(payment.qoldiq)
            
#             # Oylar sonini hisoblash (boshlang'ich to'lovsiz)
#             months_count = len([p for p in payment_schedule_data if not p['is_initial']])
            
#             # Shartnoma ma'lumotlarini tayyorlash
#             contract_data = {
#                 'id': contract.id,
#                 'contract_number': contract.contract,
#                 'status': contract.status,
#                 'pay_date': contract.pay_date,
#                 'total_price': int(contract.home.home.price * contract.home.home.field),
#                 'client_name': contract.client.full_name if contract.client else '',
#                 'home_info': f"{contract.home.building.name} - {contract.home.home.home_number}" if contract.home else ''
#             }
            
#             import json
            
#             # Frontend uchun kontekst tayyorlash
#             return render(
#                 request,
#                 "contract/list.html",
#                 {
#                     "contract": contract,
#                     "rasrochka": rasrochka,
#                     "price": contract.home.home.price * contract.home.home.field,
#                     "pk": contract_id,
#                     "payment_schedule_data": json.dumps(payment_schedule_data),
#                     "total_amount": int(total_amount),
#                     "total_paid": int(total_paid),
#                     "total_remaining": contract.residual,
#                     "months_count": months_count,
#                     "contract_data": json.dumps(contract_data),
#                     "qolgan_price": contract.residual
#                 },
#             )
    
#     messages.warning(request, "Xonadon haqida ma'lumotlar aniqlanmadi")
#     return render(
#         request,
#         "contract/list.html",
#         {
#             "contract": contract,
#             "rasrochka": Rasrochka.objects.filter(client=contract).order_by("date"),
#             "price": contract.home.home.price * contract.home.home.field,
#             "pk": contract_id
#         },
#     )
    

# def handle_custom_payment(request, contract: Client, contract_id, next_url):
#     """
#     Ixtiyoriy to'lovni qayta ishlash uchun funksiya.
#     Bu funksiya ixtiyoriy miqdordagi to'lovni qabul qiladi va
#     uni to'lanmagan oylarga taqsimlaydi.
#     """
#     # To'lov miqdorini olish
#     custom_amount_str = request.POST.get("customAmount", "0")
#     # Decimal conversion qo'shildi
#     clean_amount = custom_amount_str.replace(' ', '').replace(',', '')
#     custom_amount = Decimal(str(int(clean_amount)))

    
#     # To'lov miqdorini tekshirish
#     if custom_amount < 1:
#         messages.warning(request, "To'lov miqdori kamida 1 so'm bo'lishi kerak")
#         return redirect(f"/contract/{contract_id}/list/?next={urlencode({'next': next_url})}")
    
#     # To'lanmagan oylarni olish (sana bo'yicha tartiblangan)
#     unpaid_months = Rasrochka.objects.filter(
#         client=contract, 
#         qoldiq__gt=0
#     ).order_by("date")
    
#     # To'lanmagan oylar mavjudligini tekshirish
#     if not unpaid_months.exists():
#         messages.warning(request, "Barcha to'lovlar to'langan")
#         return redirect(f"/contract/{contract_id}/list/?next={urlencode({'next': next_url})}")
    
#     # To'lovni taqsimlash
#     remaining_amount = custom_amount
#     payments_made = 0
#     last_month_paid = None
    
#     # To'lovni oylar bo'yicha taqsimlash
#     for month in unpaid_months:
#         if remaining_amount <= 0:
#             break
        
#         # Ushbu oy uchun to'lov miqdorini hisoblash
#         amount_to_pay = min(remaining_amount, month.qoldiq)
        
#         # Oylik to'lovni yangilash
#         month.amount_paid += amount_to_pay
#         month.pay_date = datetime.now()
        
#         # Qoldiqni yangilash
#         month.qoldiq = max(0, month.amount - month.amount_paid)
#         month.save()
        
#         # Qolgan miqdorni kamaytirish
#         remaining_amount -= amount_to_pay
#         payments_made += 1
#         last_month_paid = month
    
#     # Shartnoma qoldig'ini yangilash - Decimal conversion qo'shildi
#     contract.residual -= custom_amount
    
#     # Agar qoldiq 0 bo'lsa, shartnomani tugallangan deb belgilash
#     if contract.residual <= 0:
#         contract.debt = False
#         if contract.status == 'Rasmiylashtirilgan':
#             contract.status = 'Tugallangan'
#     else:
#         contract.debt = True
        
#     contract.save()
    
#     # Muvaffaqiyatli xabar ko'rsatish
#     if payments_made > 0:
#         if last_month_paid and last_month_paid.qoldiq > 0:
#             formatted_qoldiq = "{:,}".format(int(last_month_paid.qoldiq)).replace(",", " ")
#             messages.success(
#                 request, 
#                 f"To'lov muvaffaqiyatli qabul qilindi. {last_month_paid.month}-oy uchun qolgan qarz: {formatted_qoldiq} so'm"
#             )
#         else:
#             messages.success(request, f"To'lov muvaffaqiyatli qabul qilindi")
#     else:
#         messages.warning(request, "To'lov qabul qilinmadi")
    
#     # Redirect back to payment schedule with filters
#     return redirect(f"/contract/{contract_id}/list/?next={urlencode({'next': next_url})}")   


def ClientDownload(request):
    # HTML tarkibini tayyorlash
    html_content = """
    <html>
        <head>
            <style>
            .title{
                font-size: 22px;
               text-align: center;
               border-bottom: 1px solid black;
                      font-family: "Times New Roman", Times, serif;
            }
                table {
                    width: 100%;
                    border-collapse: collapse;
                }
                th, td {
                    border: 1px solid black;
                    padding: 6px 4px;
                    text-align: center;
                    font-size: 15px;
                    font-family: "Times New Roman", Times, serif;
                }
                th {
                    background-color: #f2f2f2;
                }
                .n{
                    width: 20%
                }
                 .m{
                    width: 40%
                }
                 .i{
                    width: 40%
                }
            </style>
        </head>
        <body>
            <h2 class="title">Barcha mijzolar ro'yxati</h2>
            
            
            <table>
                <thead>
                    <tr>
                        <th class="n">N</th>
                        <th class="m">To'liq ismi</th>
                        <th class="i">Telefon raqami</th>
                        <th class="i">Qayerda eshitgan</th>
                        <th class="i">Qo'shilgan sanasi</th>
                    </tr>
                </thead>
                <tbody>
    """
    s = 1
    for row in ClientInformation.objects.all():
        html_content += f"""
                    <tr>
                        
                        <td>{s}</td>
                        <td>{row.full_name}</td>
                        <td>{row.phone or ""}\n {row.phone2 or ""}</td>
                        <td>{row.heard}</td>
                        <td>{row.created.date().strftime('%d.%m.%Y')}</td>
                    </tr>
        """
        s += 1

    html_content += """
                </tbody>
            </table>
        </body>
    </html>
    """

    # PDFni yaratish
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="mijzolar royxati.pdf"'

    pisa_status = pisa.CreatePDF(BytesIO(html_content.encode("utf-8")), dest=response)
    if pisa_status.err:
        return HttpResponse("PDF yaratishda xatolik yuz berdi", status=500)

    return response

def ClientDownloadExcel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Mijozlar"

    # Header
    headers = [
        "N",
        "To'liq ismi",
        "Telefon raqami",
        "Qayerda eshitgan",
        "Qo'shilgan sanasi"
    ]
    ws.append(headers)

    # Data
    for idx, row in enumerate(ClientInformation.objects.all(), start=1):
        ws.append([
            idx,
            row.full_name,
            f"{row.phone or ''} {row.phone2 or ''}",
            row.heard,
            row.created.date().strftime('%d.%m.%Y')
        ])

    # Set column widths for better readability
    for i, column_width in enumerate([6, 30, 25, 20, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = column_width

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="mijozlar_royxati.xlsx"'
    wb.save(response)
    return response

def number_to_words_uz(number):
    units = [
        "",
        "бир",
        "икки",
        "уч",
        "тўрт",
        "беш",
        "олти",
        "етти",
        "саккиз",
        "тўққиз",
    ]
    tens = [
        "",
        "ўн",
        "йигирма",
        "ўттиз",
        "қирқ",
        "эллик",
        "олтимиш",
        "етмиш",
        "саксон",
        "тўқсон",
    ]
    scales = ["", "минг", "миллион", "миллиард", "триллион", "квадриллион"]

    def integer_to_words(num):
        if num == 0:
            return "нол"
        words = []
        num_str = str(num)[::-1]
        groups = [num_str[i : i + 3] for i in range(0, len(num_str), 3)]

        for idx, group in enumerate(groups):
            group_word = []
            hundreds, remainder = divmod(int(group[::-1]), 100)
            tens_unit = remainder % 10
            tens_place = remainder // 10

            if hundreds > 0:
                group_word.append(units[hundreds] + " юз")

            if tens_place > 0:
                group_word.append(tens[tens_place])

            if tens_unit > 0:
                group_word.append(units[tens_unit])

            if group_word and scales[idx]:
                group_word.append(scales[idx])

            words = group_word + words

        return " ".join(words)

    integer_part = int(number)
    fractional_part = round(number % 1, 2)
    fractional_str = str(fractional_part)[2:] if fractional_part > 0 else None

    result = integer_to_words(integer_part)
    if fractional_str:
        result += f" бутун {integer_to_words(int(fractional_str))}"

    return result

def qisqartirish(full_name):
    parts = full_name.split()
    if len(parts) == 3 or len(parts) == 4:
        return f"{parts[0]} {parts[1][0].upper()}. {parts[2][0].upper()}"
    elif len(parts) == 2:
        return f"{parts[0]} {parts[1][0].upper()}."
    elif len(parts) == 1:
        return parts[0]
    return full_name


@login_required(login_url='login')
def ContractCreate(request):

    status_choices = ConsultingContract.StatusChoices.choices
    suggestion = (ConsultingContract.objects.aggregate(max_num=Max("contract_number"))["max_num"] or 0) + 1

    if request.method == "POST":
        payload, errors, client_data = _extract_consulting_contract_payload(request.POST, min_contract_number=suggestion)
        if errors:
            for error in errors:
                messages.warning(request, error)
            return redirect("contract-create")

        # Avtomatik shartnoma raqami va sanasi
        payload["contract_number"] = suggestion
        if not payload.get("contract_date"):
            payload["contract_date"] = timezone.now().date()

        # Yagona shartnoma raqami tekshiruvi
        if payload.get("contract_number") is not None and ConsultingContract.objects.filter(contract_number=payload["contract_number"]).exists():
            messages.warning(request, f"Bu shartnoma raqami ({payload['contract_number']}) allaqachon mavjud.")
            return redirect("contract-create")

        try:
            # Mijoz ma'lumotlarini yaratish yoki topish
            client_info, created = ClientInformation.objects.get_or_create(
                first_name=client_data["first_name"],
                last_name=client_data["last_name"],
                phone=client_data["phone"],
                defaults={
                    "middle_name": client_data.get("middle_name"),
                    "phone2": client_data.get("phone2"),
                    "passport_number": client_data.get("passport_number"),
                    "passport_issue_date": client_data.get("passport_issue_date"),
                    "passport_expiry_date": client_data.get("passport_expiry_date"),
                    "passport_issue_place": client_data.get("passport_issue_place"),
                    "birth_date": client_data.get("birth_date"),
                    "address": client_data.get("address"),
                    "heard": client_data.get("heard") or "Xech qayerda",
                }
            )
            
            # Mijoz ma'lumotlarini yangilash (agar o'zgargan bo'lsa)
            if not created:
                if client_data.get("middle_name"):
                    client_info.middle_name = client_data["middle_name"]
                if client_data.get("phone2"):
                    client_info.phone2 = client_data["phone2"]
                if client_data.get("passport_number"):
                    client_info.passport_number = client_data["passport_number"]
                if client_data.get("passport_issue_date"):
                    client_info.passport_issue_date = client_data["passport_issue_date"]
                if client_data.get("passport_expiry_date"):
                    client_info.passport_expiry_date = client_data["passport_expiry_date"]
                if client_data.get("passport_issue_place"):
                    client_info.passport_issue_place = client_data["passport_issue_place"]
                # birth_date ni har doim yangilash (bo'sh bo'lsa ham None sifatida saqlanadi)
                birth_date_value = client_data.get("birth_date")
                client_info.birth_date = birth_date_value.strip() if birth_date_value and birth_date_value.strip() else None
                if client_data.get("address"):
                    client_info.address = client_data["address"]
                if client_data.get("email"):
                    client_info.email = client_data["email"]
                if client_data.get("password"):
                    client_info.password = client_data["password"]
                if client_data.get("heard"):
                    client_info.heard = client_data["heard"]
                client_info.save()
            else:
                # Yangi mijoz yaratilganda email va password qo'shish
                if client_data.get("email"):
                    client_info.email = client_data["email"]
                if client_data.get("password"):
                    client_info.password = client_data["password"]
                if client_data.get("heard"):
                    client_info.heard = client_data["heard"]
                if client_data.get("passport_expiry_date"):
                    client_info.passport_expiry_date = client_data["passport_expiry_date"]
                # birth_date ni har doim yangilash (bo'sh bo'lsa ham None sifatida saqlanadi)
                birth_date_value = client_data.get("birth_date")
                client_info.birth_date = birth_date_value.strip() if birth_date_value and birth_date_value.strip() else None
                client_info.save()
            
            client_full_name = f"{client_data.get('last_name', '')} {client_data.get('first_name', '')}".strip()
            # Rasmlarni yuklab olish
            passport_images = _handle_uploaded_images(request, "passport_image", 1, contract_number=payload.get("contract_number"), client_name=client_full_name)
            visa_images = _handle_uploaded_images(request, "visa_image", 1, contract_number=payload.get("contract_number"), client_name=client_full_name)
            completed_contract_images = _handle_uploaded_images(request, "completed_contract_image", 3, contract_number=payload.get("contract_number"), client_name=client_full_name)
            
            # Shartnoma yaratish
            contract = ConsultingContract.objects.create(
                client=client_info,
                passport_images=passport_images,
                visa_images=visa_images,
                completed_contract_images=completed_contract_images,
                created_by=request.user,
                **payload
            )
            
            # Ota-onasi/farzandlari ma'lumotlarini qayta ishlash
            family_member_index = 0
            while True:
                member_first_name = request.POST.get(f"family_member_{family_member_index}_first_name", "").strip()
                member_last_name = request.POST.get(f"family_member_{family_member_index}_last_name", "").strip()
                if not member_first_name or not member_last_name:
                    break
                
                member_middle_name = request.POST.get(f"family_member_{family_member_index}_middle_name", "").strip() or None
                member_relationship = request.POST.get(f"family_member_{family_member_index}_relationship", "other")
                member_passport = request.POST.get(f"family_member_{family_member_index}_passport", "").strip() or None
                member_passport_date = request.POST.get(f"family_member_{family_member_index}_passport_date", "").strip() or None
                member_passport_expiry = request.POST.get(f"family_member_{family_member_index}_passport_expiry_date", "").strip() or None
                member_passport_place = request.POST.get(f"family_member_{family_member_index}_passport_place", "").strip() or None
                member_birth_date = request.POST.get(f"family_member_{family_member_index}_birth_date", "").strip() or None
                member_phone = request.POST.get(f"family_member_{family_member_index}_phone", "").strip() or None
                member_notes = request.POST.get(f"family_member_{family_member_index}_notes", "").strip() or None
                
                ContractFamilyMember.objects.create(
                    contract=contract,
                    first_name=member_first_name,
                    last_name=member_last_name,
                    middle_name=member_middle_name,
                    relationship=member_relationship,
                    passport_number=member_passport,
                    passport_issue_date=member_passport_date,
                    passport_expiry_date=member_passport_expiry,
                    passport_issue_place=member_passport_place,
                    birth_date=member_birth_date,
                    phone=member_phone,
                    notes=member_notes,
                )
                
                family_member_index += 1
                
        except IntegrityError as e:
            print(e)
            existing = None
            if payload.get("contract_number") is not None:
                existing = ConsultingContract.objects.filter(contract_number=payload["contract_number"]).first()
            if existing:
                messages.warning(request, f"Bu shartnoma raqami ({payload['contract_number']}) allaqachon mavjud.")
            else:
                messages.warning(request, "Shartnoma saqlashda xatolik yuz berdi.")
            return redirect("contract-create")
        except Exception as e:
            print(e)
            messages.warning(request, "Shartnoma saqlashda xatolik yuz berdi.")
            return redirect("contract-create")

        messages.success(request, "Konsalting shartnomasi muvaffaqiyatli yaratildi.")
        return redirect("contract")

    context = {
        "status_choices": status_choices,
        "suggested_contract_number": suggestion,
        "min_contract_number": suggestion,
        "today": timezone.now().date(),
    }
    return render(request, "contract/create.html", context)


@login_required(login_url='login')
def ContractEdit(request, id):

    contract = get_object_or_404(ConsultingContract, pk=id)
    status_choices = ConsultingContract.StatusChoices.choices
    min_contract_number = (ConsultingContract.objects.exclude(pk=id).aggregate(max_num=Max("contract_number"))["max_num"] or 0) + 1

    if request.method == "POST":
        payload, errors, client_data = _extract_consulting_contract_payload(request.POST, min_contract_number=min_contract_number, existing_id=id, auto_generate=False)
        if errors:
            for error in errors:
                messages.warning(request, error)
            return redirect("contract-edit", id=id)

        # Yagona shartnoma raqami tekshiruvi
        if payload.get("contract_number") is not None and ConsultingContract.objects.exclude(pk=id).filter(contract_number=payload["contract_number"]).exists():
            messages.warning(request, f"Bu shartnoma raqami ({payload['contract_number']}) allaqachon mavjud.")
            return redirect("contract-edit", id=id)

        try:
            # Mijoz ma'lumotlarini yangilash yoki yaratish
            if contract.client:
                # Mavjud mijoz ma'lumotlarini yangilash
                client_info = contract.client
                client_info.first_name = client_data["first_name"]
                client_info.last_name = client_data["last_name"]
                client_info.middle_name = client_data.get("middle_name")
                client_info.phone = client_data["phone"]
                if client_data.get("phone2"):
                    client_info.phone2 = client_data["phone2"]
                if client_data.get("passport_number"):
                    client_info.passport_number = client_data["passport_number"]
                if client_data.get("passport_issue_date"):
                    client_info.passport_issue_date = client_data["passport_issue_date"]
                if client_data.get("passport_expiry_date"):
                    client_info.passport_expiry_date = client_data["passport_expiry_date"]
                if client_data.get("passport_issue_place"):
                    client_info.passport_issue_place = client_data["passport_issue_place"]
                # birth_date ni har doim yangilash (bo'sh bo'lsa ham None sifatida saqlanadi)
                birth_date_value = client_data.get("birth_date")
                client_info.birth_date = birth_date_value.strip() if birth_date_value and birth_date_value.strip() else None
                if client_data.get("address"):
                    client_info.address = client_data["address"]
                if client_data.get("email"):
                    client_info.email = client_data["email"]
                if client_data.get("password"):
                    client_info.password = client_data["password"]
                if client_data.get("heard"):
                    client_info.heard = client_data["heard"]
                client_info.save()
            else:
                # Yangi mijoz yaratish
                client_info, created = ClientInformation.objects.get_or_create(
                    first_name=client_data["first_name"],
                    last_name=client_data["last_name"],
                    phone=client_data["phone"],
                    defaults={
                        "middle_name": client_data.get("middle_name"),
                        "phone2": client_data.get("phone2"),
                        "passport_number": client_data.get("passport_number"),
                        "passport_issue_date": client_data.get("passport_issue_date"),
                        "passport_expiry_date": client_data.get("passport_expiry_date"),
                        "passport_issue_place": client_data.get("passport_issue_place"),
                        "birth_date": client_data.get("birth_date"),
                        "address": client_data.get("address"),
                        "email": client_data.get("email"),
                        "password": client_data.get("password"),
                        "heard": client_data.get("heard") or "Xech qayerda",
                    }
                )
                if not created:
                    if client_data.get("middle_name"):
                        client_info.middle_name = client_data["middle_name"]
                    if client_data.get("phone2"):
                        client_info.phone2 = client_data["phone2"]
                    if client_data.get("passport_number"):
                        client_info.passport_number = client_data["passport_number"]
                    if client_data.get("passport_issue_date"):
                        client_info.passport_issue_date = client_data["passport_issue_date"]
                    if client_data.get("passport_expiry_date"):
                        client_info.passport_expiry_date = client_data["passport_expiry_date"]
                    if client_data.get("passport_issue_place"):
                        client_info.passport_issue_place = client_data["passport_issue_place"]
                    # birth_date ni har doim yangilash (bo'sh bo'lsa ham None sifatida saqlanadi)
                    birth_date_value = client_data.get("birth_date")
                    client_info.birth_date = birth_date_value.strip() if birth_date_value and birth_date_value.strip() else None
                    if client_data.get("address"):
                        client_info.address = client_data["address"]
                    if client_data.get("email"):
                        client_info.email = client_data["email"]
                    if client_data.get("password"):
                        client_info.password = client_data["password"]
                    if client_data.get("heard"):
                        client_info.heard = client_data["heard"]
                    client_info.save()
            
            # Rasmlarni yuklab olish (mavjud rasmlarni saqlab qolish)
            existing_passport_images = list(contract.passport_images or [])
            existing_completed_images = list(contract.completed_contract_images or [])
            existing_visa_images = list(contract.visa_images or [])
            
            # Eski rasmlarni o'chirish (agar belgilangan bo'lsa)
            delete_passport_indices = []
            for idx_str in request.POST.getlist("delete_passport_image"):
                try:
                    idx = int(idx_str)
                    delete_passport_indices.append(idx)
                except (ValueError, TypeError):
                    pass
            
            delete_completed_indices = []
            for idx_str in request.POST.getlist("delete_completed_contract_image"):
                try:
                    idx = int(idx_str)
                    delete_completed_indices.append(idx)
                except (ValueError, TypeError):
                    pass
            delete_visa_indices = []
            for idx_str in request.POST.getlist("delete_visa_image"):
                try:
                    idx = int(idx_str)
                    delete_visa_indices.append(idx)
                except (ValueError, TypeError):
                    pass
            
            # O'chirilgan rasmlarni olib tashlash (indekslar bo'yicha)
            passport_images = [img for idx, img in enumerate(existing_passport_images) if idx not in delete_passport_indices]
            completed_contract_images = [img for idx, img in enumerate(existing_completed_images) if idx not in delete_completed_indices]
            visa_images = [img for idx, img in enumerate(existing_visa_images) if idx not in delete_visa_indices]
            
            client_full_name = f"{client_data.get('last_name', '')} {client_data.get('first_name', '')}".strip()
            # Yangi rasmlarni qo'shish (cheklovlar bilan)
            new_passport_images = _handle_uploaded_images(request, "passport_image", 1, contract_number=contract.contract_number, client_name=client_full_name)
            new_visa_images = _handle_uploaded_images(request, "visa_image", 1, contract_number=contract.contract_number, client_name=client_full_name)
            new_completed_images = _handle_uploaded_images(request, "completed_contract_image", 3, contract_number=contract.contract_number, client_name=client_full_name)
            
            # Mavjud rasmlar sonini hisoblab, qolgan joyga yangilarini qo'shish
            remaining_passport_slots = 1 - len(passport_images)
            remaining_completed_slots = 3 - len(completed_contract_images)
            
            if remaining_passport_slots > 0:
                passport_images.extend(new_passport_images[:remaining_passport_slots])
            
            # Visa rasmlari: agar yangi yuklansa, eski o'rniga qo'yamiz
            if new_visa_images:
                visa_images = new_visa_images[:1]

            if remaining_completed_slots > 0:
                completed_contract_images.extend(new_completed_images[:remaining_completed_slots])
            
            # Shartnoma ma'lumotlarini yangilash
            for field, value in payload.items():
                setattr(contract, field, value)
            contract.client = client_info
            contract.passport_images = passport_images
            contract.visa_images = visa_images
            contract.completed_contract_images = completed_contract_images
            contract.save()
            
            # Ota-onasi/farzandlari ma'lumotlarini yangilash
            # Avval mavjudlarini o'chirish
            contract.family_members.all().delete()
            
            # Yangilarini qo'shish
            family_member_index = 0
            while True:
                member_first_name = request.POST.get(f"family_member_{family_member_index}_first_name", "").strip()
                member_last_name = request.POST.get(f"family_member_{family_member_index}_last_name", "").strip()
                if not member_first_name or not member_last_name:
                    break
                
                member_middle_name = request.POST.get(f"family_member_{family_member_index}_middle_name", "").strip() or None
                member_relationship = request.POST.get(f"family_member_{family_member_index}_relationship", "other")
                member_passport = request.POST.get(f"family_member_{family_member_index}_passport", "").strip() or None
                member_passport_date = request.POST.get(f"family_member_{family_member_index}_passport_date", "").strip() or None
                member_passport_expiry = request.POST.get(f"family_member_{family_member_index}_passport_expiry_date", "").strip() or None
                member_passport_place = request.POST.get(f"family_member_{family_member_index}_passport_place", "").strip() or None
                member_birth_date = request.POST.get(f"family_member_{family_member_index}_birth_date", "").strip() or None
                member_phone = request.POST.get(f"family_member_{family_member_index}_phone", "").strip() or None
                member_notes = request.POST.get(f"family_member_{family_member_index}_notes", "").strip() or None
                
                ContractFamilyMember.objects.create(
                    contract=contract,
                    first_name=member_first_name,
                    last_name=member_last_name,
                    middle_name=member_middle_name,
                    relationship=member_relationship,
                    passport_number=member_passport,
                    passport_issue_date=member_passport_date,
                    passport_expiry_date=member_passport_expiry,
                    passport_issue_place=member_passport_place,
                    birth_date=member_birth_date,
                    phone=member_phone,
                    notes=member_notes,
                )
                
                family_member_index += 1
                
        except IntegrityError:
            existing = None
            if payload.get("contract_number") is not None:
                existing = ConsultingContract.objects.exclude(pk=id).filter(contract_number=payload["contract_number"]).first()
            if existing:
                messages.warning(request, f"Bu shartnoma raqami ({payload['contract_number']}) allaqachon mavjud.")
            else:
                messages.warning(request, "Shartnoma saqlashda xatolik yuz berdi.")
            return redirect("contract-edit", id=id)
        except Exception as e:
            print(e)
            messages.warning(request, "Shartnoma saqlashda xatolik yuz berdi.")
            return redirect("contract-edit", id=id)

        messages.success(request, "Shartnoma ma'lumotlari yangilandi.")
        return redirect("contract")

    # Mijoz ma'lumotlarini context ga qo'shish
    client_info = contract.client if contract.client else None
    family_members = contract.family_members.all() if hasattr(contract, 'family_members') else []

    context = {
        "contract": contract,
        "client_info": client_info,
        "family_members": family_members,
        "status_choices": status_choices,
        "min_contract_number": min_contract_number,
    }
    return render(request, "contract/edit.html", context)


@login_required(login_url='login')
@require_POST
def ContractDelete(request, id):
    contract = get_object_or_404(ConsultingContract, pk=id)
    contract.delete()
    messages.success(request, "Shartnoma muvaffaqiyatli o'chirildi.")
    return redirect("contract")


@login_required(login_url='login')
def ContractCreatePDF(request, id):
    contract = get_object_or_404(ConsultingContract, pk=id)

    month_name = [
        "Yanvar", "Fevral", "Mart", "Aprel", "May", "Iyun",
        "Iyul", "Avgust", "Sentabr", "Oktabr", "Noyabr", "Dekabr"
    ]
    month = month_name[contract.contract_date.month - 1]

    total_price = contract.total_service_fee
    initial_paid = contract.initial_payment_amount
    remaining_balance = max(total_price - initial_paid, 0)

    foiz = 0
    if total_price:
        foiz = (initial_paid / total_price) * 100
    foiz_formatted = f"{foiz:.2f}".rstrip('0').rstrip('.') if foiz else "0"

    month_names_uz = {
        1: "yanvar", 2: "fevral", 3: "mart", 4: "aprel", 5: "may", 6: "iyun",
        7: "iyul", 8: "avgust", 9: "sentabr", 10: "oktabr", 11: "noyabr", 12: "dekabr"
    }


    html_content = render_to_string(
        "shart.html",
        {
            'rounded_home_price': int(total_price),
            "pk": contract.contract_number,
            "contract": contract,
            "month": month,
            "price": int(total_price),
            "mijoz_tolagan": int(initial_paid),
            "qoldiq": int(remaining_balance),
            "foiz": foiz_formatted,
            "total_price": 0,
            "down_payment": 0,
            "remaining_balance": remaining_balance
        },
    )

    font_config = FontConfiguration()
    pdf = HTML(string=html_content, base_url=request.build_absolute_uri()).write_pdf(font_config=font_config)

    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="consulting-contract-{id}.pdf"'
    return response

    
    
def LoginPage(request):
    if request.user.is_authenticated:
        return redirect("home")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        
        if not username or not password:
            return render(
                request=request,
                template_name="login.html",
                context={
                    "error": "Login va parol kiritilishi shart",
                },
            )
            
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            return redirect("home")
        else:
            return render(
                request=request,
                template_name="login.html",
                context={
                    "error": "Login yoki parol noto'g'ri",
                    "userr": username,
                },
            )
    return render(request=request, template_name="login.html")

def LogoutPage(request):
    logout(request)
    return redirect("login")

@login_required(login_url='login')
def NotificationsPage(request):
    """
    Konsalting shartnomalari bo'yicha bildirishnomalar:
    - Overdue: status completed/cancelled emas va remaining_amount > 0
    - Upcoming: status draft va remaining_amount > 0
    """

    overdue_contracts = ConsultingContract.objects.filter(
        amount_paid__lt=F("total_service_fee"),
        status__in=[
            ConsultingContract.StatusChoices.SUBMITTED,
            ConsultingContract.StatusChoices.PREPARATION,
        ],
    ).select_related("client")

    upcoming_contracts = ConsultingContract.objects.filter(
        amount_paid__lt=F("total_service_fee"),
        status__in=[ConsultingContract.StatusChoices.DRAFT],
    ).select_related("client")

    context = {
        'overdue_contracts': overdue_contracts,
        'upcoming_contracts': upcoming_contracts,
        'overdue_count': overdue_contracts.count(),
        'upcoming_count': upcoming_contracts.count(),
        'total_notifications': overdue_contracts.count() + upcoming_contracts.count(),
    }

    return render(request, 'notifications.html', context)

