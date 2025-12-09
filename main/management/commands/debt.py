from django.core.management.base import BaseCommand
import time
from main.models import Client, ClientInformation
from main.views import send_sms
import asyncio
from datetime import datetime, timedelta
import re


def normalize_phone(phon):
    if not phon:
        return None
    # Float ko'rinishida kelsa, butun qismini olish
    if isinstance(phon, float):
        phon = str(int(phon))
    phone = str(phon)
    digits = re.sub(r'\D', '', phone)
    # Agar raqam 9 tadan ko'p bo'lsa, oxirgi 9 tasini olish
    if len(digits) > 9:
        digits = digits[-9:]
    if len(digits) == 9:
        return '+998' + digits
    elif len(digits) == 12 and digits.startswith('998'):
        return '+' + digits
    else:
        return None

class Command(BaseCommand):
    def handle(self, *args, **kwargs):
        import pandas as pd

        # Excel faylingiz nomi
        file_path = '/root/qurilish/main/management/commands/mijozlar.xlsx'  # faylni shu nom bilan saqlang

        from openpyxl import load_workbook

        # Excel faylni ochish
        workbook = load_workbook(filename=file_path)
        sheet = workbook.active

        # Har bir qatordagi A (telefon raqam) va B (ism) ustunlarini o‘qish
        for row in sheet.iter_rows(min_row=2):  # 1-qator sarlavha deb faraz qilamiz
            telefon = row[0].value  # A ustun
            ism = row[1].value      # B ustun
            
            if telefon and isinstance(telefon, (int, str)):
                telefon_str = str(telefon)
                if telefon_str.isdigit() and len(telefon_str) == 9:
                    try:
                        c = ClientInformation.objects.create(full_name=ism, phone=normalize_phone(telefon_str))
                        print(telefon_str, True)
                    except:
                        print(telefon_str, False)
                        continue